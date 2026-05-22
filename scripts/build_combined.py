#!/usr/bin/env python3
"""Build a combined xlsx file from the JSON data sources.

The xlsx is a read-only view of the framework data. The JSON files in data/
are the canonical source of truth. Edits should be made to the JSON files,
not to the spreadsheet.

Run from the repository root: python scripts/build_combined.py [output.xlsx]
"""

import json
import sys
from pathlib import Path

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter
except ImportError:
    print("Required package missing. Install with:")
    print("  pip install openpyxl")
    sys.exit(2)


REPO_ROOT = Path(__file__).resolve().parent.parent
DATA_DIR = REPO_ROOT / "data"
DEFAULT_OUTPUT = REPO_ROOT / "build" / "levels-of-interoperability.xlsx"


def load_json(path):
    with open(path, "r", encoding="utf-8") as f:
        return json.load(f)


def format_list(value):
    """Format list-valued cells as comma-separated strings."""
    if value is None:
        return ""
    if isinstance(value, list):
        return ", ".join(str(v) for v in value)
    return str(value)


def write_sheet(wb, sheet_name, columns, rows, freeze=True):
    """Write a sheet with header row and data rows.

    columns is a list of (field_name, display_label) tuples.
    rows is a list of dicts.
    """
    ws = wb.create_sheet(sheet_name)

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="2C3E50", end_color="2C3E50", fill_type="solid")
    header_align = Alignment(horizontal="left", vertical="top", wrap_text=True)
    data_align = Alignment(horizontal="left", vertical="top", wrap_text=True)

    # Header
    for col_idx, (_, label) in enumerate(columns, 1):
        cell = ws.cell(row=1, column=col_idx, value=label)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align

    # Data
    for row_idx, row in enumerate(rows, 2):
        for col_idx, (field, _) in enumerate(columns, 1):
            value = row.get(field)
            if isinstance(value, bool):
                cell_value = "yes" if value else "no"
            elif isinstance(value, list):
                cell_value = format_list(value)
            elif value is None:
                cell_value = ""
            else:
                cell_value = str(value)
            cell = ws.cell(row=row_idx, column=col_idx, value=cell_value)
            cell.alignment = data_align

    # Column widths: rough heuristic based on field type
    width_map = {
        "id": 8,
        "order": 8,
        "name": 30,
        "category": 18,
        "kind": 15,
        "description": 50,
    }
    for col_idx, (field, _) in enumerate(columns, 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width_map.get(field, 35)

    # Freeze the header row and the id column
    if freeze:
        ws.freeze_panes = "C2"

    # Reasonable row heights for readability
    for row_idx in range(2, len(rows) + 2):
        ws.row_dimensions[row_idx].height = 60


def build_levels_sheet(wb, data):
    columns = [
        ("id", "ID"),
        ("order", "Order"),
        ("name", "Treatment level"),
        ("description", "Description"),
        ("keyCharacteristics", "Key characteristics"),
        ("scopeApplies", "Scope applies"),
        ("accessibilityConsideration", "Accessibility consideration"),
        ("licensingConsideration", "Licensing consideration"),
        ("populationCompletenessApplies", "Population completeness applies"),
        ("populationCompletenessNotes", "Population completeness notes"),
        ("typicallyRequires", "Typically requires"),
        ("typicallyRequiresNotes", "Typically requires notes"),
        ("branchOrJumpNotes", "Branch or jump notes"),
        ("whatItEnables", "What it enables"),
        ("fairContributionLocal", "FAIR (local)"),
        ("fairContributionExternal", "FAIR (external)"),
        ("notes", "Notes"),
    ]
    entries = sorted(data["entries"], key=lambda e: e["order"])
    write_sheet(wb, "Treatment Levels", columns, entries)


def build_resources_sheet(wb, data):
    columns = [
        ("id", "ID"),
        ("order", "Order"),
        ("name", "Resource type"),
        ("description", "Description"),
        ("typicalForm", "Typical form"),
        ("commonlyPresent", "Commonly present"),
        ("realisticCeiling", "Realistic ceiling"),
        ("realisticCeilingLevels", "Realistic ceiling levels"),
        ("scopeNotes", "Scope notes"),
        ("accessibilityNotes", "Accessibility notes"),
        ("licensingNotes", "Licensing notes"),
        ("notes", "Notes"),
    ]
    entries = sorted(data["entries"], key=lambda e: e["order"])
    write_sheet(wb, "Resources", columns, entries)


def build_tools_sheet(wb, data):
    columns = [
        ("id", "ID"),
        ("order", "Order"),
        ("name", "Name"),
        ("category", "Category"),
        ("description", "Description"),
        ("handlesResources", "Handles resources"),
        ("handlesAllResources", "Handles all resources"),
        ("handlesResourcesNotes", "Handles resources notes"),
        ("supportsLevels", "Supports levels"),
        ("supportsLevelsNotes", "Supports levels notes"),
        ("enablesTransitions", "Enables transitions"),
        ("enablesTransitionsNotes", "Enables transitions notes"),
        ("currentStatus", "Current status"),
        ("currentStatusNotes", "Current status notes"),
        ("ownersAndInitiatives", "Owners and initiatives"),
        ("dependsOn", "Depends on"),
        ("notes", "Notes"),
    ]
    entries = sorted(data["entries"], key=lambda e: e["order"])
    write_sheet(wb, "Tools and Services", columns, entries)


def build_transitions_sheet(wb, data):
    columns = [
        ("id", "ID"),
        ("order", "Order"),
        ("kind", "Kind"),
        ("from", "From"),
        ("fromNotes", "From notes"),
        ("to", "To"),
        ("toNotes", "To notes"),
        ("description", "Description"),
        ("resourcesAffected", "Resources affected"),
        ("resourcesAffectedAllResources", "All resources affected"),
        ("resourcesAffectedNotes", "Resources affected notes"),
        ("workRequired", "Work required"),
        ("toolsAndServicesInvolved", "Tools and services involved"),
        ("notes", "Notes"),
    ]
    entries = sorted(data["entries"], key=lambda e: e["order"])
    write_sheet(wb, "Transitions", columns, entries)


def build_vocabularies_sheet(wb, data):
    """Vocabularies are nested; flatten for the spreadsheet view."""
    ws = wb.create_sheet("Vocabularies")

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="2C3E50", end_color="2C3E50", fill_type="solid")
    header_align = Alignment(horizontal="left", vertical="top", wrap_text=True)
    data_align = Alignment(horizontal="left", vertical="top", wrap_text=True)

    columns = [
        ("vocabulary", "Vocabulary"),
        ("vocab_description", "Vocabulary description"),
        ("id", "Term ID"),
        ("value", "Value"),
        ("label", "Label"),
        ("description", "Description"),
    ]

    for col_idx, (_, label) in enumerate(columns, 1):
        cell = ws.cell(row=1, column=col_idx, value=label)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align

    row_idx = 2
    for vocab in data["vocabularies"]:
        vocab_name = vocab["name"]
        vocab_desc = vocab.get("description", "")
        for term in vocab["terms"]:
            values = {
                "vocabulary": vocab_name,
                "vocab_description": vocab_desc,
                "id": term["id"],
                "value": term["value"],
                "label": term.get("label", ""),
                "description": term.get("description", ""),
            }
            for col_idx, (field, _) in enumerate(columns, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=values[field])
                cell.alignment = data_align
            row_idx += 1

    width_map = {
        "vocabulary": 20,
        "vocab_description": 40,
        "id": 8,
        "value": 25,
        "label": 25,
        "description": 50,
    }
    for col_idx, (field, _) in enumerate(columns, 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width_map.get(field, 30)

    ws.freeze_panes = "C2"


def build_overview_sheet(wb, levels, resources, tools, transitions):
    """A summary sheet that briefly describes the framework and links to the other sheets."""
    ws = wb.create_sheet("Overview", 0)

    title_font = Font(bold=True, size=18)
    heading_font = Font(bold=True, size=12)

    ws["A1"] = "Levels of Interoperability for Heritage Science"
    ws["A1"].font = title_font

    ws["A3"] = f"Version: {levels.get('version', 'unknown')}"
    ws["A4"] = f"Last updated: {levels.get('lastUpdated', 'unknown')}"

    ws["A6"] = "Contents"
    ws["A6"].font = heading_font

    contents = [
        ("Treatment Levels", f"{len(levels['entries'])} levels", "How resources can be organised and described, from tacit and unstructured through to integrated digital constructs."),
        ("Resources", f"{len(resources['entries'])} resource types", "Types of material researchers and institutions may hold."),
        ("Tools and Services", f"{len(tools['entries'])} entries", "Tools, services, infrastructure, and reference resources that support the framework."),
        ("Transitions", f"{len(transitions['entries'])} transitions", "Movements between levels and cross-cutting changes."),
        ("Vocabularies", "Multiple", "Controlled vocabularies used by the framework."),
    ]

    for i, (sheet_name, count, desc) in enumerate(contents, 7):
        ws.cell(row=i, column=1, value=sheet_name).font = Font(bold=True)
        ws.cell(row=i, column=2, value=count)
        ws.cell(row=i, column=3, value=desc).alignment = Alignment(wrap_text=True, vertical="top")

    ws["A14"] = "Note"
    ws["A14"].font = heading_font
    ws["A15"] = "This spreadsheet is generated from JSON data files in the data/ directory of the repository. Edits should be made to the JSON files, not to this spreadsheet, as the spreadsheet is rebuilt on every release."
    ws["A15"].alignment = Alignment(wrap_text=True, vertical="top")
    ws.row_dimensions[15].height = 60

    ws.column_dimensions["A"].width = 25
    ws.column_dimensions["B"].width = 22
    ws.column_dimensions["C"].width = 70


def main():
    output_path = Path(sys.argv[1]) if len(sys.argv) > 1 else DEFAULT_OUTPUT
    output_path.parent.mkdir(parents=True, exist_ok=True)

    levels = load_json(DATA_DIR / "treatment-levels.json")
    resources = load_json(DATA_DIR / "resources.json")
    tools = load_json(DATA_DIR / "tools-and-services.json")
    transitions = load_json(DATA_DIR / "transitions.json")
    vocabularies = load_json(DATA_DIR / "vocabularies.json")

    wb = Workbook()
    # Remove the default sheet
    wb.remove(wb.active)

    build_levels_sheet(wb, levels)
    build_resources_sheet(wb, resources)
    build_tools_sheet(wb, tools)
    build_transitions_sheet(wb, transitions)
    build_vocabularies_sheet(wb, vocabularies)
    build_overview_sheet(wb, levels, resources, tools, transitions)

    wb.save(output_path)
    print(f"Wrote {output_path}")


if __name__ == "__main__":
    main()
